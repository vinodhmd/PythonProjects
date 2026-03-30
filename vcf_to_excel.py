"""
vcf_to_excel.py
---------------
Parses a VCF (vCard 2.1 / 3.0) file and exports all contacts to an Excel file.
Handles folded lines, multiple phone numbers, email addresses, and organisations.
Photos (base64 blobs) are intentionally skipped.
"""

import re
import sys
import os
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Installing openpyxl...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter


# ── helpers ──────────────────────────────────────────────────────────────────

def decode_qp(text: str) -> str:
    """Decode Quoted-Printable encoded string."""
    try:
        import quopri
        return quopri.decodestring(text.encode()).decode("utf-8", errors="replace")
    except Exception:
        return text


def clean_value(value: str, encoding: str = "") -> str:
    """Strip and optionally decode a vCard field value."""
    value = value.strip()
    if "QUOTED-PRINTABLE" in encoding.upper():
        value = decode_qp(value)
    return value


def parse_vcf(vcf_path: str) -> list[dict]:
    """
    Parse a VCF file and return a list of contact dicts.
    Each dict contains:
        name, first_name, last_name, display_name,
        phone_mobile, phone_home, phone_work, phone_other,
        email_home, email_work, email_other,
        org, title, address, note, birthday, url
    """
    contacts = []
    current: dict | None = None
    in_photo = False          # True while consuming a folded PHOTO block
    pending_key = ""          # key for the next folded-continuation line
    pending_val = ""          # accumulated value

    PHOTO_RE = re.compile(r"^PHOTO\s*[;:]", re.IGNORECASE)

    def flush_field():
        nonlocal pending_key, pending_val
        if current is not None and pending_key:
            _store_field(current, pending_key, pending_val)
        pending_key = ""
        pending_val = ""

    with open(vcf_path, "r", encoding="utf-8", errors="replace") as fh:
        for raw_line in fh:
            line = raw_line.rstrip("\r\n")

            # ── folded-line continuation ──
            if line.startswith((" ", "\t")):
                if in_photo:
                    continue   # discard photo data
                if pending_key:
                    pending_val += line[1:]  # strip the leading whitespace
                continue

            # Non-continuation → flush the previous field
            flush_field()
            in_photo = False

            # ── PHOTO lines: start skipping ──
            if PHOTO_RE.match(line):
                in_photo = True
                continue

            # ── BEGIN:VCARD ──
            if line.upper().startswith("BEGIN:VCARD"):
                current = {
                    "name": "", "first_name": "", "last_name": "", "display_name": "",
                    "phone_mobile": [], "phone_home": [], "phone_work": [], "phone_other": [],
                    "email_home": [], "email_work": [], "email_other": [],
                    "org": "", "title": "", "address": "", "note": "", "birthday": "", "url": "",
                    "nickname": ""
                }
                continue

            # ── END:VCARD ──
            if line.upper().startswith("END:VCARD"):
                if current is not None:
                    contacts.append(current)
                    current = None
                continue

            # ── Normal field line ──
            if current is None:
                continue

            # Split key and value at the first colon NOT inside a group prefix
            colon_pos = _find_colon(line)
            if colon_pos < 0:
                continue

            pending_key = line[:colon_pos].strip()
            pending_val = line[colon_pos + 1:]  # rest may continue on next lines

    # Flush last field in case file doesn't end with END:VCARD
    flush_field()

    return contacts


def _find_colon(line: str) -> int:
    """Return the position of the first ':' that is not inside a parameter."""
    # Simple approach: first colon after any semicolon params
    return line.find(":")


def _store_field(contact: dict, key: str, raw_value: str):
    """Interpret a vCard field and store it in the contact dict."""
    key_upper = key.upper()
    parts = key_upper.split(";")
    base_key = parts[0].strip()
    params = [p.strip() for p in parts[1:]]

    # Encoding detection
    encoding = next((p for p in params if "ENCODING" in p or "QUOTED" in p), "")
    value = clean_value(raw_value, encoding)

    # ── N (structured name) ──
    if base_key == "N":
        n_parts = value.split(";")
        contact["last_name"] = n_parts[0].strip() if len(n_parts) > 0 else ""
        contact["first_name"] = n_parts[1].strip() if len(n_parts) > 1 else ""
        contact["name"] = f"{contact['first_name']} {contact['last_name']}".strip()
        return

    # ── FN (formatted / display name) ──
    if base_key == "FN":
        contact["display_name"] = value
        if not contact["name"]:
            contact["name"] = value
        return

    # ── TEL ──
    if base_key == "TEL":
        # Normalise the phone number
        phone = re.sub(r"[\s\-\(\)]", "", value)
        param_str = " ".join(params)
        if any(t in param_str for t in ("CELL", "MOBILE")):
            contact["phone_mobile"].append(phone)
        elif "HOME" in param_str:
            contact["phone_home"].append(phone)
        elif "WORK" in param_str:
            contact["phone_work"].append(phone)
        else:
            contact["phone_other"].append(phone)
        return

    # ── EMAIL ──
    if base_key == "EMAIL":
        param_str = " ".join(params)
        if "HOME" in param_str:
            contact["email_home"].append(value)
        elif "WORK" in param_str:
            contact["email_work"].append(value)
        else:
            contact["email_other"].append(value)
        return

    # ── ORG ──
    if base_key == "ORG":
        contact["org"] = value.replace(";", " | ").strip(" |")
        return

    # ── TITLE ──
    if base_key == "TITLE":
        contact["title"] = value
        return

    # ── ADR ──
    if base_key == "ADR":
        adr_parts = value.split(";")
        full_adr = ", ".join(p.strip() for p in adr_parts if p.strip())
        if contact["address"]:
            contact["address"] += " | " + full_adr
        else:
            contact["address"] = full_adr
        return

    # ── NOTE ──
    if base_key == "NOTE":
        contact["note"] = value
        return

    # ── BDAY (birthday) ──
    if base_key == "BDAY":
        contact["birthday"] = value
        return

    # ── URL ──
    if base_key == "URL":
        contact["url"] = value
        return

    # ── NICKNAME ──
    if base_key == "NICKNAME":
        contact["nickname"] = value
        return


def contacts_to_excel(contacts: list[dict], output_path: str):
    """Write the contacts list to a nicely formatted Excel file."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Contacts"

    # ── Column definitions ──
    columns = [
        ("S.No",            10),
        ("Display Name",    28),
        ("First Name",      20),
        ("Last Name",       20),
        ("Mobile",          18),
        ("Phone (Home)",    18),
        ("Phone (Work)",    18),
        ("Phone (Other)",   18),
        ("Email (Home)",    30),
        ("Email (Work)",    30),
        ("Email (Other)",   30),
        ("Organisation",    28),
        ("Title",           20),
        ("Nickname",        18),
        ("Birthday",        14),
        ("Address",         40),
        ("URL",             30),
        ("Note",            40),
    ]

    # ── Styles ──
    HEADER_FILL  = PatternFill("solid", fgColor="1F3864")   # dark navy
    HEADER_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    ALT_FILL     = PatternFill("solid", fgColor="EBF0FA")   # light blue-grey
    BODY_FONT    = Font(name="Calibri", size=10)
    CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT_ALIGN   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thin_side    = Side(style="thin", color="B0B8C8")
    BORDER       = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    # ── Header row ──
    ws.row_dimensions[1].height = 32
    for col_idx, (header, width) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border    = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ── Freeze header ──
    ws.freeze_panes = "A2"

    # ── Auto-filter ──
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}1"

    # ── Data rows ──
    for row_idx, c in enumerate(contacts, start=2):
        fill = ALT_FILL if row_idx % 2 == 0 else PatternFill()

        def w(col, value):
            cell = ws.cell(row=row_idx, column=col, value=value or "")
            cell.font      = BODY_FONT
            cell.fill      = fill
            cell.border    = BORDER
            cell.alignment = LEFT_ALIGN if col > 1 else CENTER_ALIGN

        w(1,  row_idx - 1)
        w(2,  c["display_name"] or c["name"])
        w(3,  c["first_name"])
        w(4,  c["last_name"])
        w(5,  " / ".join(c["phone_mobile"]))
        w(6,  " / ".join(c["phone_home"]))
        w(7,  " / ".join(c["phone_work"]))
        w(8,  " / ".join(c["phone_other"]))
        w(9,  " / ".join(c["email_home"]))
        w(10, " / ".join(c["email_work"]))
        w(11, " / ".join(c["email_other"]))
        w(12, c["org"])
        w(13, c["title"])
        w(14, c["nickname"])
        w(15, c["birthday"])
        w(16, c["address"])
        w(17, c["url"])
        w(18, c["note"])

        ws.row_dimensions[row_idx].height = 18

    # ── Summary sheet ──
    ws_sum = wb.create_sheet("Summary")
    ws_sum["A1"] = "VCF Contact Export Summary"
    ws_sum["A1"].font = Font(name="Calibri", bold=True, size=14)
    ws_sum["A3"] = "Source file:"
    ws_sum["B3"] = os.path.basename(vcf_path)
    ws_sum["A4"] = "Extracted on:"
    ws_sum["B4"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws_sum["A5"] = "Total contacts:"
    ws_sum["B5"] = len(contacts)
    ws_sum.column_dimensions["A"].width = 20
    ws_sum.column_dimensions["B"].width = 35

    wb.save(output_path)
    print(f"\n[OK] Saved {len(contacts):,} contacts -> {output_path}")


# ── main ──────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    vcf_dir = Path(r"c:\Vinodh\Work\Pythonbase\PythonProjects")

    # Accept an optional command-line argument
    if len(sys.argv) > 1:
        vcf_path = sys.argv[1]
    else:
        vcf_path = str(vcf_dir / "vcards_20250808_201257.vcf")

    if not os.path.exists(vcf_path):
        print(f"ERROR: VCF file not found: {vcf_path}")
        sys.exit(1)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = str(vcf_dir / f"contacts_export_{timestamp}.xlsx")

    print(f"[>>] Reading:  {vcf_path}")
    print("[..] Parsing contacts (skipping photos)...")

    contacts = parse_vcf(vcf_path)
    print(f"[>>] Found {len(contacts):,} contacts.")

    print("[>>] Writing Excel...")
    contacts_to_excel(contacts, output_path)
